from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import datetime, timedelta
import json

from shop.models import Category, SubCategory, Type, Product, ProductImage, ProductSpecification, Brand, HeroSlide
from orders.models import Order, OrderItem, Delivery
from .forms import CategoryForm, SubCategoryForm, TypeForm, ProductForm, OrderStatusForm, DeliveryForm, HeroSlideForm
from .excel_import import ExcelImporter
import os
from django.conf import settings


# ==================== Authentification ====================

def admin_login(request):
    if request.user.is_authenticated:
        return redirect('admin_panel:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_panel:dashboard')
        else:
            messages.error(request, 'Identifiants invalides ou accès non autorisé.')
    
    return render(request, 'admin_panel/login.html')


@login_required
def admin_logout(request):
    logout(request)
    return redirect('admin_panel:login')


# ==================== Dashboard ====================

@login_required
def dashboard(request):
    # Filtres par date
    period = request.GET.get('period', 'all')  # all, today, week, month
    
    # Définir les dates selon la période
    now = timezone.now()
    date_from = None
    date_to = None
    
    # Si une période rapide est sélectionnée, utiliser les calculs de dates
    # IMPORTANT: Ne PAS utiliser les paramètres date_from/date_to de l'URL si period != 'all'
    if period == 'today':
        date_from = now.date()
        date_to = now.date()
    elif period == 'week':
        date_from = (now - timedelta(days=7)).date()
        date_to = now.date()
    elif period == 'month':
        date_from = (now - timedelta(days=30)).date()
        date_to = now.date()
    elif period == 'all':
        # Utiliser les dates manuelles uniquement si period='all'
        date_from_str = request.GET.get('date_from', '').strip()
        date_to_str = request.GET.get('date_to', '').strip()
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
    
    # Filtrer les commandes selon la période - utiliser des datetimes au lieu de dates
    orders_query = Order.objects.all()
    
    if date_from:
        # Convertir la date en datetime avec début de journée (00:00:00)
        from datetime import datetime as dt
        datetime_from = timezone.make_aware(dt.combine(date_from, dt.min.time()))
        orders_query = orders_query.filter(created_at__gte=datetime_from)
    if date_to:
        # Convertir la date en datetime avec fin de journée (23:59:59)
        from datetime import datetime as dt
        datetime_to = timezone.make_aware(dt.combine(date_to, dt.max.time()))
        orders_query = orders_query.filter(created_at__lte=datetime_to)
    
    # Statistiques globales
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    total_users = User.objects.filter(is_staff=True).count()
    
    # Statistiques des commandes
    total_orders = orders_query.count()
    pending_orders = orders_query.filter(status='pending').count()
    confirmed_orders = orders_query.filter(status='confirmed').count()
    # Compter les livraisons réellement livrées (Delivery.status='delivered') pour les commandes filtrées
    delivered_orders = Delivery.objects.filter(order__in=orders_query, status='delivered').count()
    cancelled_orders = orders_query.filter(status='cancelled').count()
    
    # Revenus
    total_revenue = orders_query.filter(status__in=['confirmed', 'delivered']).aggregate(
        total=Sum('total'))['total'] or 0
    pending_revenue = orders_query.filter(status='pending').aggregate(
        total=Sum('total'))['total'] or 0
    
    # Statistiques par jour (pour le graphique)
    daily_stats_query = orders_query.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id'),
        revenue=Sum('total')
    ).order_by('-date')[:30]
    
    # Convertir en liste pour le template et JSON
    daily_stats = []
    for stat in daily_stats_query:
        # Vérifier que la date existe
        if stat['date'] is not None:
            daily_stats.append({
                'date': stat['date'].strftime('%d/%m'),
                'count': stat['count'],
                'revenue': float(stat['revenue'] or 0)
            })
    daily_stats.reverse()  # Ordre chronologique pour le graphique
    
    # Dernières commandes
    recent_orders = orders_query.select_related('customer').order_by('-created_at')[:10]
    
    # Produits les plus vendus dans la période
    from orders.models import OrderItem
    top_products = OrderItem.objects.filter(
        order__in=orders_query
    ).values(
        'product__name', 'product__reference'
    ).annotate(
        quantity=Sum('quantity')
    ).order_by('-quantity')[:5]
    
    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_users': total_users,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'confirmed_orders': confirmed_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'daily_stats': json.dumps(daily_stats) if daily_stats else '[]',
        'has_stats': len(daily_stats) > 0,
        'date_from': date_from,
        'date_to': date_to,
        'date_from_str': date_from.strftime('%Y-%m-%d') if date_from else '',
        'date_to_str': date_to.strftime('%Y-%m-%d') if date_to else '',
        'period': period,
    }
    return render(request, 'admin_panel/dashboard.html', context)


# ==================== Catégories ====================

@login_required
def category_list(request):
    categories = Category.objects.all()
    
    # Filtre par recherche
    search = request.GET.get('search', '')
    if search:
        categories = categories.filter(Q(name__icontains=search) | Q(description__icontains=search))
    
    # Filtre par statut
    status = request.GET.get('status', '')
    if status == 'active':
        categories = categories.filter(is_active=True)
    elif status == 'inactive':
        categories = categories.filter(is_active=False)
    
    categories = categories.order_by('order', 'name')
    return render(request, 'admin_panel/category_list.html', {
        'categories': categories,
        'search': search,
        'status': status,
    })


@login_required
def category_add(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie ajoutée avec succès.')
            return redirect('admin_panel:category_list')
    else:
        form = CategoryForm()
    return render(request, 'admin_panel/category_form.html', {'form': form, 'title': 'Ajouter une catégorie'})


@login_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        # Vérifier si on veut supprimer l'image
        if request.POST.get('clear_image'):
            category.image.delete(save=False)
            category.image = None
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie modifiée avec succès.')
            return redirect('admin_panel:category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'admin_panel/category_form.html', {
        'form': form, 
        'title': 'Modifier la catégorie',
        'category': category
    })


@login_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Catégorie supprimée avec succès.')
        return redirect('admin_panel:category_list')
    return render(request, 'admin_panel/category_confirm_delete.html', {'category': category})


# ==================== Sous-catégories ====================

@login_required
def subcategory_list(request):
    subcategories = SubCategory.objects.select_related('category')
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    # Filtre par recherche
    search = request.GET.get('search', '')
    if search:
        subcategories = subcategories.filter(
            Q(name__icontains=search) | 
            Q(description__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Filtre par catégorie
    category_id = request.GET.get('category', '')
    if category_id:
        subcategories = subcategories.filter(category_id=category_id)
    
    # Filtre par statut
    status = request.GET.get('status', '')
    if status == 'active':
        subcategories = subcategories.filter(is_active=True)
    elif status == 'inactive':
        subcategories = subcategories.filter(is_active=False)
    
    # Filtre par affichage page d'accueil
    homepage = request.GET.get('homepage', '')
    if homepage == 'yes':
        subcategories = subcategories.filter(show_on_homepage=True)
    elif homepage == 'no':
        subcategories = subcategories.filter(show_on_homepage=False)
    
    # Filtre par sous-catégorie essentielle
    essential = request.GET.get('essential', '')
    if essential == 'yes':
        subcategories = subcategories.filter(is_essential=True)
    elif essential == 'no':
        subcategories = subcategories.filter(is_essential=False)
    
    subcategories = subcategories.order_by('category__name', 'order', 'name')
    return render(request, 'admin_panel/subcategory_list.html', {
        'subcategories': subcategories,
        'categories': categories,
        'search': search,
        'category_id': category_id,
        'status': status,
        'homepage': homepage,
        'essential': essential,
    })


@login_required
def subcategory_add(request):
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sous-catégorie ajoutée avec succès.')
            return redirect('admin_panel:subcategory_list')
    else:
        form = SubCategoryForm()
    return render(request, 'admin_panel/subcategory_form.html', {'form': form, 'title': 'Ajouter une sous-catégorie'})


@login_required
def subcategory_edit(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        # Vérifier si on veut supprimer l'image
        if request.POST.get('clear_image'):
            subcategory.image.delete(save=False)
            subcategory.image = None
        form = SubCategoryForm(request.POST, request.FILES, instance=subcategory)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sous-catégorie modifiée avec succès.')
            return redirect('admin_panel:subcategory_list')
    else:
        form = SubCategoryForm(instance=subcategory)
    return render(request, 'admin_panel/subcategory_form.html', {
        'form': form, 
        'title': 'Modifier la sous-catégorie',
        'subcategory': subcategory
    })


@login_required
def subcategory_delete(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        subcategory.delete()
        messages.success(request, 'Sous-catégorie supprimée avec succès.')
        return redirect('admin_panel:subcategory_list')
    return render(request, 'admin_panel/subcategory_confirm_delete.html', {'subcategory': subcategory})


# ==================== Types ====================

@login_required
def type_list(request):
    types = Type.objects.select_related('subcategory', 'subcategory__category', 'brand')
    categories = Category.objects.filter(is_active=True).order_by('name')
    subcategories = SubCategory.objects.filter(is_active=True).order_by('name')
    brands = Brand.objects.filter(is_active=True).order_by('name')
    
    # Filtre par recherche
    search = request.GET.get('search', '')
    if search:
        types = types.filter(
            Q(name__icontains=search) | 
            Q(description__icontains=search) |
            Q(subcategory__name__icontains=search) |
            Q(subcategory__category__name__icontains=search) |
            Q(brand__name__icontains=search)
        )
    
    # Filtre par marque
    brand_id = request.GET.get('brand', '')
    if brand_id:
        types = types.filter(brand_id=brand_id)
    
    # Filtre par catégorie
    category_id = request.GET.get('category', '')
    if category_id:
        types = types.filter(subcategory__category_id=category_id)
    
    # Filtre par sous-catégorie
    subcategory_id = request.GET.get('subcategory', '')
    if subcategory_id:
        types = types.filter(subcategory_id=subcategory_id)
    
    # Filtre par statut
    status = request.GET.get('status', '')
    if status == 'active':
        types = types.filter(is_active=True)
    elif status == 'inactive':
        types = types.filter(is_active=False)
    
    types = types.order_by('subcategory__name', 'order', 'name')
    return render(request, 'admin_panel/type_list.html', {
        'types': types,
        'categories': categories,
        'subcategories': subcategories,
        'brands': brands,
        'search': search,
        'brand_id': brand_id,
        'category_id': category_id,
        'subcategory_id': subcategory_id,
        'status': status,
    })


@login_required
def type_add(request):
    if request.method == 'POST':
        form = TypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Type ajouté avec succès.')
            return redirect('admin_panel:type_list')
    else:
        form = TypeForm()
    return render(request, 'admin_panel/type_form.html', {'form': form, 'title': 'Ajouter un type'})


@login_required
def type_edit(request, pk):
    type_obj = get_object_or_404(Type, pk=pk)
    if request.method == 'POST':
        form = TypeForm(request.POST, instance=type_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Type modifié avec succès.')
            return redirect('admin_panel:type_list')
    else:
        form = TypeForm(instance=type_obj)
    return render(request, 'admin_panel/type_form.html', {'form': form, 'title': 'Modifier le type'})


@login_required
def type_delete(request, pk):
    type_obj = get_object_or_404(Type, pk=pk)
    if request.method == 'POST':
        type_obj.delete()
        messages.success(request, 'Type supprimé avec succès.')
        return redirect('admin_panel:type_list')
    return render(request, 'admin_panel/type_confirm_delete.html', {'type': type_obj})


# ==================== Marques ====================

@login_required
def brand_list(request):
    brands = Brand.objects.all()
    
    # Filtre par recherche
    search = request.GET.get('search', '')
    if search:
        brands = brands.filter(Q(name__icontains=search) | Q(description__icontains=search))
    
    # Filtre par statut
    status = request.GET.get('status', '')
    if status == 'active':
        brands = brands.filter(is_active=True)
    elif status == 'inactive':
        brands = brands.filter(is_active=False)
    
    brands = brands.order_by('order', 'name')
    return render(request, 'admin_panel/brand_list.html', {
        'brands': brands,
        'search': search,
        'status': status,
    })


@login_required
def brand_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        logo = request.FILES.get('logo')
        logo_url = request.POST.get('logo_url', '')
        description = request.POST.get('description', '')
        order = request.POST.get('order', 0)
        is_active = request.POST.get('is_active') == 'on'
        
        brand = Brand.objects.create(
            name=name,
            logo=logo,
            logo_url=logo_url,
            description=description,
            order=order,
            is_active=is_active
        )
        messages.success(request, 'Marque ajoutée avec succès.')
        return redirect('admin_panel:brand_list')
    
    return render(request, 'admin_panel/brand_form.html', {'title': 'Ajouter une marque'})


@login_required
def brand_edit(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.name = request.POST.get('name')
        
        # Vérifier si on veut supprimer le logo
        if request.POST.get('clear_logo'):
            brand.logo.delete(save=False)
            brand.logo = None
        elif 'logo' in request.FILES:
            brand.logo = request.FILES.get('logo')
        
        brand.logo_url = request.POST.get('logo_url', '')
        brand.description = request.POST.get('description', '')
        brand.order = request.POST.get('order', 0)
        brand.is_active = request.POST.get('is_active') == 'on'
        brand.save()
        
        messages.success(request, 'Marque modifiée avec succès.')
        return redirect('admin_panel:brand_list')
    
    return render(request, 'admin_panel/brand_form.html', {'brand': brand, 'title': 'Modifier la marque'})


@login_required
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, 'Marque supprimée avec succès.')
        return redirect('admin_panel:brand_list')
    return render(request, 'admin_panel/brand_confirm_delete.html', {'brand': brand})


# ==================== Produits ====================

@login_required
def product_list(request):
    products = Product.objects.select_related('category', 'subcategory', 'brand', 'type').prefetch_related('images').order_by('-created_at')
    
    # Récupérer tous les éléments pour les filtres
    categories = Category.objects.filter(is_active=True).order_by('name')
    subcategories = SubCategory.objects.filter(is_active=True).order_by('name')
    brands = Brand.objects.filter(is_active=True).order_by('name')
    types = Type.objects.filter(is_active=True).order_by('name')
    
    # Filtres
    category_id = request.GET.get('category')
    subcategory_id = request.GET.get('subcategory')
    brand_id = request.GET.get('brand')
    type_id = request.GET.get('type')
    search = request.GET.get('search')
    status = request.GET.get('status')
    stock_status = request.GET.get('stock_status')
    is_bestseller = request.GET.get('is_bestseller')
    is_featured = request.GET.get('is_featured')
    is_new = request.GET.get('is_new')
    
    # Appliquer les filtres
    if category_id:
        products = products.filter(category_id=category_id)
        # Filtrer les sous-catégories par catégorie sélectionnée
        subcategories = subcategories.filter(category_id=category_id)
    
    if subcategory_id:
        products = products.filter(subcategory_id=subcategory_id)
    
    if brand_id:
        products = products.filter(brand_id=brand_id)
    
    if type_id:
        products = products.filter(type_id=type_id)
    
    if search:
        products = products.filter(
            Q(name__icontains=search) | 
            Q(reference__icontains=search) | 
            Q(description__icontains=search) |
            Q(brand__name__icontains=search)
        )
    
    if status:
        products = products.filter(status=status)
    
    # Filtre de stock
    if stock_status == 'in_stock':
        products = products.filter(quantity__gt=0)
    elif stock_status == 'out_of_stock':
        products = products.filter(quantity=0)
    elif stock_status == 'low_stock':
        products = products.filter(quantity__gt=0, quantity__lte=5)
    
    # Filtres booléens
    if is_bestseller == 'yes':
        products = products.filter(is_bestseller=True)
    elif is_bestseller == 'no':
        products = products.filter(is_bestseller=False)
    
    if is_featured == 'yes':
        products = products.filter(is_featured=True)
    elif is_featured == 'no':
        products = products.filter(is_featured=False)
    
    if is_new == 'yes':
        products = products.filter(is_new=True)
    elif is_new == 'no':
        products = products.filter(is_new=False)
    
    # Compter les résultats
    total_count = products.count()
    
    return render(request, 'admin_panel/product_list.html', {
        'products': products,
        'categories': categories,
        'subcategories': subcategories,
        'brands': brands,
        'types': types,
        'total_count': total_count,
        # Garder les valeurs des filtres
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
        'selected_brand': brand_id,
        'selected_type': type_id,
        'selected_status': status,
        'selected_stock_status': stock_status,
        'selected_is_bestseller': is_bestseller,
        'selected_is_featured': is_featured,
        'selected_is_new': is_new,
        'search_query': search,
    })


@login_required
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            
            # Gérer les images multiples
            images = request.FILES.getlist('product_images')
            main_image_index = request.POST.get('main_image_index', '0')
            
            for idx, image in enumerate(images):
                is_main = (str(idx) == main_image_index)
                ProductImage.objects.create(
                    product=product,
                    image=image,
                    is_main=is_main,
                    order=idx
                )
            
            # Gérer les caractéristiques
            spec_keys = request.POST.getlist('spec_key[]')
            spec_values = request.POST.getlist('spec_value[]')
            
            for idx, (key, value) in enumerate(zip(spec_keys, spec_values)):
                if key.strip() and value.strip():  # Ignorer les champs vides
                    ProductSpecification.objects.create(
                        product=product,
                        key=key.strip(),
                        value=value.strip(),
                        order=idx
                    )
            
            messages.success(request, 'Produit ajouté avec succès.')
            return redirect('admin_panel:product_list')
    else:
        form = ProductForm()
    
    return render(request, 'admin_panel/product_form.html', {
        'form': form,
        'title': 'Ajouter un produit',
    })


@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save()
            
            # Gérer les nouvelles images
            images = request.FILES.getlist('product_images')
            if images:
                # Récupérer l'index de l'image principale parmi les nouvelles images
                main_image_index = request.POST.get('main_image_index', '')
                
                # Calculer l'ordre de départ (après les images existantes)
                existing_count = product.images.count()
                
                for idx, image in enumerate(images):
                    is_main = (str(idx) == main_image_index)
                    # Si cette nouvelle image est principale, désélectionner les autres
                    if is_main:
                        product.images.update(is_main=False)
                    
                    ProductImage.objects.create(
                        product=product,
                        image=image,
                        is_main=is_main,
                        order=existing_count + idx
                    )
            
            # Gérer la mise à jour de l'image principale existante
            existing_main_id = request.POST.get('existing_main_image')
            if existing_main_id:
                product.images.update(is_main=False)
                product.images.filter(id=existing_main_id).update(is_main=True)
            
            # Gérer les caractéristiques
            # D'abord, supprimer les anciennes caractéristiques
            product.specifications.all().delete()
            
            # Ajouter les nouvelles caractéristiques
            spec_keys = request.POST.getlist('spec_key[]')
            spec_values = request.POST.getlist('spec_value[]')
            
            for idx, (key, value) in enumerate(zip(spec_keys, spec_values)):
                if key.strip() and value.strip():
                    ProductSpecification.objects.create(
                        product=product,
                        key=key.strip(),
                        value=value.strip(),
                        order=idx
                    )
            
            messages.success(request, 'Produit modifié avec succès.')
            return redirect('admin_panel:product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'admin_panel/product_form.html', {
        'form': form, 
        'title': 'Modifier le produit',
        'product': product,
    })


@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Produit supprimé avec succès.')
        return redirect('admin_panel:product_list')
    return render(request, 'admin_panel/product_confirm_delete.html', {'product': product})


@login_required
def product_image_delete(request, pk):
    """Supprimer une image de produit"""
    from django.http import JsonResponse
    
    if request.method == 'POST':
        try:
            image = get_object_or_404(ProductImage, pk=pk)
            product = image.product
            
            # Si c'était l'image principale, définir une autre image comme principale
            if image.is_main and product.images.count() > 1:
                next_image = product.images.exclude(pk=pk).first()
                if next_image:
                    next_image.is_main = True
                    next_image.save()
            
            image.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# ==================== Commandes ====================

@login_required
def order_list(request):
    orders = Order.objects.select_related('customer').order_by('-created_at')
    
    # Filtres par statut
    status = request.GET.get('status')
    if status:
        orders = orders.filter(status=status)
    
    # Filtres par date
    period = request.GET.get('period', '')
    date_from = None
    date_to = None
    
    now = timezone.now()
    
    if period == 'today':
        date_from = now.date()
        date_to = now.date()
    elif period == 'week':
        date_from = (now - timedelta(days=7)).date()
        date_to = now.date()
    elif period == 'month':
        date_from = (now - timedelta(days=30)).date()
        date_to = now.date()
    elif period == 'all':
        # Dates manuelles
        date_from_str = request.GET.get('date_from', '').strip()
        date_to_str = request.GET.get('date_to', '').strip()
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
    
    # Appliquer les filtres de date
    if date_from:
        from datetime import datetime as dt
        datetime_from = timezone.make_aware(dt.combine(date_from, dt.min.time()))
        orders = orders.filter(created_at__gte=datetime_from)
    
    if date_to:
        from datetime import datetime as dt
        datetime_to = timezone.make_aware(dt.combine(date_to, dt.max.time()))
        orders = orders.filter(created_at__lte=datetime_to)
    
    # Préparer les dates pour le template
    date_from_str = date_from.strftime('%Y-%m-%d') if date_from else ''
    date_to_str = date_to.strftime('%Y-%m-%d') if date_to else ''
    
    return render(request, 'admin_panel/order_list.html', {
        'orders': orders,
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'date_from_str': date_from_str,
        'date_to_str': date_to_str,
    })


@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order_items = order.items.select_related('product').all()
    
    return render(request, 'admin_panel/order_detail.html', {
        'order': order,
        'order_items': order_items
    })


@login_required
def order_confirm(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        try:
            # Verifier si le stock a deja ete deduit
            if order.stock_deducted:
                messages.warning(request, f'La commande {order.order_number} a deja ete confirmee et le stock a deja ete deduit.')
                return redirect('admin_panel:order_detail', pk=pk)
            
            # Verifier le stock disponible pour tous les produits
            insufficient_stock = []
            for item in order.items.all():
                if item.product.quantity < item.quantity:
                    insufficient_stock.append(f"{item.product_name} (Stock disponible: {item.product.quantity}, Quantite demandee: {item.quantity})")
            
            if insufficient_stock:
                messages.error(request, f'Stock insuffisant pour: {", ".join(insufficient_stock)}')
                return redirect('admin_panel:order_detail', pk=pk)
            
            # Deduire les quantites du stock
            for item in order.items.all():
                product = item.product
                old_quantity = product.quantity
                product.quantity -= item.quantity
                product.save(update_fields=['quantity'])
                # Log pour debug
                print(f"Stock mis a jour pour {product.name}: {old_quantity} -> {product.quantity}")
            
            # Mettre a jour le statut de la commande et marquer le stock comme deduit
            order.status = 'confirmed'
            order.confirmed_at = timezone.now()
            order.stock_deducted = True
            order.save()
            
            # Creer une livraison si elle n'existe pas
            if not hasattr(order, 'delivery'):
                Delivery.objects.create(order=order)
            
            messages.success(request, f'Commande {order.order_number} confirmee. Le stock a ete mis a jour.')
            return redirect('admin_panel:order_detail', pk=pk)
        except Exception as e:
            import traceback
            print(f"Erreur order_confirm: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'Erreur lors de la confirmation: {str(e)}')
            return redirect('admin_panel:order_detail', pk=pk)
    return render(request, 'admin_panel/order_confirm.html', {'order': order})


@login_required
def order_cancel(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        try:
            # Si le stock a ete deduit, le restaurer
            if order.stock_deducted:
                for item in order.items.all():
                    product = item.product
                    old_quantity = product.quantity
                    product.quantity += item.quantity
                    product.save(update_fields=['quantity'])
                    # Log pour debug
                    print(f"Stock restaure pour {product.name}: {old_quantity} -> {product.quantity}")
                
                order.stock_deducted = False
                messages.success(request, f'Commande {order.order_number} annulee. Le stock a ete restaure.')
            else:
                messages.success(request, f'Commande {order.order_number} annulee.')
            
            order.status = 'cancelled'
            order.save()
            return redirect('admin_panel:order_detail', pk=pk)
        except Exception as e:
            import traceback
            print(f"Erreur order_cancel: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'Erreur lors de l annulation: {str(e)}')
            return redirect('admin_panel:order_detail', pk=pk)
    return render(request, 'admin_panel/order_cancel.html', {'order': order})


# ==================== Livraisons ====================

@login_required
def delivery_list(request):
    deliveries = Delivery.objects.select_related('order', 'order__customer').order_by('-created_at')
    
    # Filtres par statut
    status = request.GET.get('status')
    if status:
        deliveries = deliveries.filter(status=status)
    
    # Filtres par date
    period = request.GET.get('period', '')
    date_from = None
    date_to = None
    
    now = timezone.now()
    
    if period == 'today':
        date_from = now.date()
        date_to = now.date()
    elif period == 'week':
        date_from = (now - timedelta(days=7)).date()
        date_to = now.date()
    elif period == 'month':
        date_from = (now - timedelta(days=30)).date()
        date_to = now.date()
    elif period == 'all':
        # Dates manuelles
        date_from_str = request.GET.get('date_from', '').strip()
        date_to_str = request.GET.get('date_to', '').strip()
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
    
    # Appliquer les filtres de date
    if date_from:
        from datetime import datetime as dt
        datetime_from = timezone.make_aware(dt.combine(date_from, dt.min.time()))
        deliveries = deliveries.filter(created_at__gte=datetime_from)
    
    if date_to:
        from datetime import datetime as dt
        datetime_to = timezone.make_aware(dt.combine(date_to, dt.max.time()))
        deliveries = deliveries.filter(created_at__lte=datetime_to)
    
    # Préparer les dates pour le template
    date_from_str = date_from.strftime('%Y-%m-%d') if date_from else ''
    date_to_str = date_to.strftime('%Y-%m-%d') if date_to else ''
    
    return render(request, 'admin_panel/delivery_list.html', {
        'deliveries': deliveries,
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'date_from_str': date_from_str,
        'date_to_str': date_to_str,
    })


@login_required
def delivery_detail(request, pk):
    delivery = get_object_or_404(Delivery, pk=pk)
    return render(request, 'admin_panel/delivery_detail.html', {'delivery': delivery})


@login_required
def delivery_update(request, pk):
    delivery = get_object_or_404(Delivery, pk=pk)
    if request.method == 'POST':
        form = DeliveryForm(request.POST, instance=delivery)
        if form.is_valid():
            form.save()
            messages.success(request, 'Livraison mise à jour avec succès.')
            return redirect('admin_panel:delivery_detail', pk=pk)
    else:
        form = DeliveryForm(instance=delivery)
    return render(request, 'admin_panel/delivery_form.html', {'form': form, 'delivery': delivery})


# ==================== Gestion des Utilisateurs ====================

@login_required
def user_list(request):
    """Liste tous les utilisateurs admin"""
    users = User.objects.filter(is_staff=True).order_by('-date_joined')
    return render(request, 'admin_panel/user_list.html', {'users': users})


@login_required
def user_add(request):
    """Ajouter un nouvel utilisateur admin"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        is_superuser = request.POST.get('is_superuser') == 'on'
        
        # Validation
        if not username or not password:
            messages.error(request, 'Le nom d\'utilisateur et le mot de passe sont requis.')
            return render(request, 'admin_panel/user_form.html', {'title': 'Ajouter un utilisateur'})
        
        if password != password_confirm:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
            return render(request, 'admin_panel/user_form.html', {'title': 'Ajouter un utilisateur'})
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Ce nom d\'utilisateur existe déjà.')
            return render(request, 'admin_panel/user_form.html', {'title': 'Ajouter un utilisateur'})
        
        # Créer l'utilisateur
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )
        user.is_staff = True
        user.is_superuser = is_superuser
        user.save()
        
        messages.success(request, f'Utilisateur {username} créé avec succès.')
        return redirect('admin_panel:user_list')
    
    return render(request, 'admin_panel/user_form.html', {'title': 'Ajouter un utilisateur'})


@login_required
def user_edit(request, pk):
    """Modifier un utilisateur admin"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.is_superuser = request.POST.get('is_superuser') == 'on'
        user.is_active = request.POST.get('is_active') == 'on'
        
        # Changer le mot de passe seulement si fourni
        new_password = request.POST.get('password')
        if new_password:
            password_confirm = request.POST.get('password_confirm')
            if new_password == password_confirm:
                user.set_password(new_password)
            else:
                messages.error(request, 'Les mots de passe ne correspondent pas.')
                return render(request, 'admin_panel/user_form.html', {
                    'title': 'Modifier l\'utilisateur',
                    'user_obj': user
                })
        
        user.save()
        messages.success(request, f'Utilisateur {user.username} modifié avec succès.')
        return redirect('admin_panel:user_list')
    
    return render(request, 'admin_panel/user_form.html', {
        'title': 'Modifier l\'utilisateur',
        'user_obj': user
    })


@login_required
def user_delete(request, pk):
    """Supprimer un utilisateur admin"""
    user = get_object_or_404(User, pk=pk)
    
    # Empêcher la suppression de son propre compte
    if user.id == request.user.id:
        messages.error(request, 'Vous ne pouvez pas supprimer votre propre compte.')
        return redirect('admin_panel:user_list')
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'Utilisateur {username} supprimé avec succès.')
        return redirect('admin_panel:user_list')
    
    return render(request, 'admin_panel/user_confirm_delete.html', {'user_obj': user})


# ==================== AJAX Endpoints ====================

from django.http import JsonResponse

@login_required
def get_subcategories_by_category(request):
    """Retourne les sous-catégories d'une catégorie donnée"""
    category_id = request.GET.get('category_id')
    if category_id:
        subcategories = SubCategory.objects.filter(
            category_id=category_id,
            is_active=True
        ).values('id', 'name').order_by('order', 'name')
        return JsonResponse(list(subcategories), safe=False)
    return JsonResponse([], safe=False)


@login_required
def get_types_by_subcategory(request):
    """Retourne les types/modèles d'une sous-catégorie donnée"""
    subcategory_id = request.GET.get('subcategory_id')
    if subcategory_id:
        types = Type.objects.filter(
            subcategory_id=subcategory_id,
            is_active=True
        ).values('id', 'name').order_by('order', 'name')
        return JsonResponse(list(types), safe=False)
    return JsonResponse([], safe=False)


@login_required
def get_types_by_brand(request):
    """Retourne les types/modèles d'une marque donnée"""
    brand_id = request.GET.get('brand_id')
    if brand_id:
        types = Type.objects.filter(
            brand_id=brand_id,
            is_active=True
        ).values('id', 'name', 'subcategory_id').order_by('order', 'name')
        return JsonResponse(list(types), safe=False)
    return JsonResponse([], safe=False)


@login_required
def get_types_filtered(request):
    """Retourne les types/modèles filtrés par sous-catégorie ET/OU marque"""
    subcategory_id = request.GET.get('subcategory_id')
    brand_id = request.GET.get('brand_id')
    
    # Commencer avec tous les types actifs
    types = Type.objects.filter(is_active=True)
    
    # Appliquer les filtres si fournis
    if subcategory_id:
        types = types.filter(subcategory_id=subcategory_id)
    
    if brand_id:
        types = types.filter(brand_id=brand_id)
    
    # Retourner la liste
    types_list = types.values('id', 'name', 'subcategory_id', 'brand_id').order_by('order', 'name')
    return JsonResponse(list(types_list), safe=False)


# ==================== Hero Slides ====================

@login_required
def hero_slide_list(request):
    """Liste des slides hero"""
    slides = HeroSlide.objects.all().select_related('category', 'subcategory', 'product').order_by('order', '-created_at')
    
    context = {
        'slides': slides,
        'page_title': 'Gestion des Slides Hero',
    }
    return render(request, 'admin_panel/hero_slide_list.html', context)


@login_required
def hero_slide_add(request):
    """Ajouter un nouveau slide hero"""
    if request.method == 'POST':
        form = HeroSlideForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                slide = form.save()
                messages.success(request, 'Le slide hero a été ajouté avec succès.')
                return redirect('admin_panel:hero_slide_list')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'ajout du slide: {str(e)}')
        else:
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire.')
    else:
        form = HeroSlideForm()
    
    context = {
        'form': form,
        'page_title': 'Ajouter un Slide Hero',
        'action': 'Ajouter',
    }
    return render(request, 'admin_panel/hero_slide_form.html', context)


@login_required
def hero_slide_edit(request, pk):
    """Modifier un slide hero"""
    slide = get_object_or_404(HeroSlide, pk=pk)
    
    if request.method == 'POST':
        form = HeroSlideForm(request.POST, request.FILES, instance=slide)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Le slide hero a été modifié avec succès.')
                return redirect('admin_panel:hero_slide_list')
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification du slide: {str(e)}')
        else:
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire.')
    else:
        form = HeroSlideForm(instance=slide)
    
    context = {
        'form': form,
        'slide': slide,
        'page_title': 'Modifier un Slide Hero',
        'action': 'Modifier',
    }
    return render(request, 'admin_panel/hero_slide_form.html', context)


@login_required
def hero_slide_delete(request, pk):
    """Supprimer un slide hero"""
    slide = get_object_or_404(HeroSlide, pk=pk)
    
    if request.method == 'POST':
        slide.delete()
        messages.success(request, 'Le slide hero a été supprimé avec succès.')
        return redirect('admin_panel:hero_slide_list')
    
    context = {
        'slide': slide,
        'page_title': 'Confirmer la suppression',
    }
    return render(request, 'admin_panel/hero_slide_confirm_delete.html', context)


# ==================== Importation Excel ====================

@login_required
def product_import(request):
    """Page d'importation de produits depuis Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Vérifier l'extension du fichier
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Veuillez uploader un fichier Excel valide (.xlsx ou .xls)')
            return redirect('admin_panel:product_import')
        
        # Sauvegarder temporairement le fichier
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_path = tmp_file.name
        
        try:
            # Importer les produits
            importer = ExcelImporter()
            result = importer.import_from_excel(tmp_path)
            
            if result['success']:
                # Message de succes detaille
                success_msg = f"[OK] Importation terminee avec succes!\n"
                success_msg += f"• {result['created']} produits créés\n"
                success_msg += f"• {result['skipped']} produits ignorés (doublons ou données manquantes)\n"
                
                if result['created_brands']:
                    success_msg += f"• {len(result['created_brands'])} nouvelles marques créées: {', '.join(result['created_brands'][:5])}"
                    if len(result['created_brands']) > 5:
                        success_msg += f" et {len(result['created_brands']) - 5} autres"
                    success_msg += "\n"
                
                if result['created_types']:
                    success_msg += f"• {len(result['created_types'])} nouveaux types créés: {', '.join(result['created_types'][:5])}"
                    if len(result['created_types']) > 5:
                        success_msg += f" et {len(result['created_types']) - 5} autres"
                    success_msg += "\n"
                
                if result['created_collections']:
                    success_msg += f"• {len(result['created_collections'])} nouvelles collections créées: {', '.join(result['created_collections'])}\n"
                
                messages.success(request, success_msg)
                
                # Afficher les erreurs s'il y en a
                if result['errors']:
                    error_msg = "[ATTENTION] Erreurs rencontrees:\n" + "\n".join(result['errors'][:10])
                    if len(result['errors']) > 10:
                        error_msg += f"\n... et {len(result['errors']) - 10} autres erreurs"
                    messages.warning(request, error_msg)
            else:
                messages.error(request, f"[ERREUR] {result['error']}")
        
        except Exception as e:
            messages.error(request, f"[ERREUR] Erreur lors de l'importation: {str(e)}")
        
        finally:
            # Supprimer le fichier temporaire
            try:
                os.unlink(tmp_path)
            except:
                pass
        
        return redirect('admin_panel:product_import')
    
    # Statistiques actuelles
    stats = {
        'products': Product.objects.count(),
        'categories': Category.objects.count(),
        'subcategories': SubCategory.objects.count(),
        'brands': Brand.objects.count(),
        'types': Type.objects.count(),
    }
    
    return render(request, 'admin_panel/product_import.html', {'stats': stats})


# ==================== Importation des Images ====================



@login_required
def product_images_import(request):
    """Page d'importation des images de produits depuis un dossier"""
    if request.method == 'POST':
        images_path = request.POST.get('images_path', '').strip()
        
        if not images_path:
            messages.error(request, 'Veuillez fournir le chemin du dossier d\'images.')
            return redirect('admin_panel:product_images_import')
        
        # Vérifier que le chemin existe
        if not os.path.exists(images_path):
            messages.error(request, f'Le chemin "{images_path}" n\'existe pas.')
            return redirect('admin_panel:product_images_import')
        
        # Vérifier que c'est un dossier
        if not os.path.isdir(images_path):
            messages.error(request, f'"{images_path}" n\'est pas un dossier valide.')
            return redirect('admin_panel:product_images_import')
        
        # Importer le script d'importation des images
        import sys
        from pathlib import Path
        from PIL import Image
        import shutil
        from django.db.models.functions import Lower
        
        # Extensions d'images autorisées
        IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
        
        def normalize_name(name):
            """Normalise un nom pour la comparaison (minuscules, espaces supprimés)"""
            return name.lower().strip()
        
        def extract_reference_from_folder_name(folder_name):
            """Extrait le numéro de référence du nom du dossier
            
            Exemples:
            - "Carte mère ASUS ROG STRIX Z690-A GAMING WIFI D4" -> "Z690-A"
            - "Processeur Intel Core i7-12700K" -> "i7-12700K"
            - "MSI RTX 4090 GAMING X TRIO 24G" -> "RTX 4090" ou "4090"
            
            La fonction cherche les patterns courants de références:
            - Références avec tirets (ex: i7-12700K, RTX-4090)
            - Références alphanumériques (ex: Z690A, RTX4090)
            - Nombres seuls si précédés d'une marque connue
            """
            import re
            
            # Patterns de références courants
            patterns = [
                r'\b([A-Z0-9]+-[A-Z0-9-]+)\b',  # Format avec tirets: i7-12700K, RTX-4090
                r'\b(RTX\s*\d{4}\s*[A-Z]*|GTX\s*\d{4}\s*[A-Z]*)\b',  # Cartes graphiques NVIDIA
                r'\b(RX\s*\d{4}\s*[A-Z]*)\b',  # Cartes graphiques AMD
                r'\b([iI][3579]-\d{4,5}[A-Z]{0,2})\b',  # Processeurs Intel
                r'\b(Ryzen\s*[3579]\s*\d{4}[A-Z]{0,2})\b',  # Processeurs AMD Ryzen
                r'\b([A-Z]\d{3,4}[A-Z]*-[A-Z0-9]+)\b',  # Format type Z690-A, B550-F
                r'\b([A-Z]{2,}\d{3,})\b',  # Format alphanumérique: RTX4090, Z690A
            ]
            
            for pattern in patterns:
                match = re.search(pattern, folder_name, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
            
            return None
        
        def find_product_by_reference(folder_name):
            """Trouve un produit par son numéro de référence extrait du nom du dossier
            
            1. Extrait la référence du nom du dossier
            2. Cherche dans la base de données en comparant avec le champ 'reference'
            3. Si pas trouvé, cherche dans le nom du produit
            """
            # Extraire la référence du nom du dossier
            reference = extract_reference_from_folder_name(folder_name)
            
            if not reference:
                # Fallback: chercher par nom complet
                normalized_search = normalize_name(folder_name)
                products = Product.objects.annotate(name_lower=Lower('name'))
                for product in products:
                    if normalize_name(product.name) == normalized_search:
                        return product, None
                return None, None
            
            # Nettoyer la référence
            reference_clean = reference.upper().strip()
            
            # 1. Chercher d'abord dans le champ reference (correspondance exacte)
            product = Product.objects.filter(reference__iexact=reference_clean).first()
            if product:
                return product, reference
            
            # 2. Chercher dans le champ reference (contient)
            product = Product.objects.filter(reference__icontains=reference_clean).first()
            if product:
                return product, reference
            
            # 3. Chercher dans le nom du produit (contient la référence)
            product = Product.objects.filter(name__icontains=reference).first()
            if product:
                return product, reference
            
            # 4. Fallback: chercher par nom complet du dossier
            normalized_search = normalize_name(folder_name)
            products = Product.objects.annotate(name_lower=Lower('name'))
            for product in products:
                if normalize_name(product.name) == normalized_search:
                    return product, reference
            
            return None, reference
        
        def get_image_files(directory):
            """Récupère tous les fichiers images d'un dossier"""
            if not os.path.exists(directory):
                return []
            
            image_files = []
            for file in os.listdir(directory):
                file_path = os.path.join(directory, file)
                if os.path.isfile(file_path):
                    ext = os.path.splitext(file)[1].lower()
                    if ext in IMAGE_EXTENSIONS:
                        image_files.append(file_path)
            
            return sorted(image_files)
        
        def copy_image_to_media(source_path, product, is_main=False):
            """Copie une image vers le dossier media et crée l'entrée en base"""
            try:
                # Vérifier que le fichier existe
                if not os.path.exists(source_path):
                    return False, f"Fichier introuvable: {source_path}"
                
                # Générer le nom du fichier de destination
                filename = os.path.basename(source_path)
                destination_subdir = 'products/gallery/'
                
                # Chemin relatif pour Django (depuis media/)
                relative_path = os.path.join(destination_subdir, filename)
                
                # Vérifier si cette image existe déjà pour ce produit (basé sur le nom du fichier)
                existing_image = ProductImage.objects.filter(
                    product=product,
                    image__icontains=filename
                ).first()
                
                if existing_image:
                    return False, f"Image déjà existante (ignorée): {filename}"
                
                # Ouvrir et vérifier l'image
                try:
                    with Image.open(source_path) as img:
                        img.verify()
                except Exception as e:
                    return False, f"Image corrompue {os.path.basename(source_path)}: {e}"
                
                # Créer le dossier de destination s'il n'existe pas
                media_root = os.path.join(settings.BASE_DIR, 'media', destination_subdir)
                os.makedirs(media_root, exist_ok=True)
                
                # Chemin absolu pour la copie
                destination_path = os.path.join(settings.BASE_DIR, 'media', relative_path)
                
                # Copier le fichier seulement s'il n'existe pas déjà
                if not os.path.exists(destination_path):
                    shutil.copy2(source_path, destination_path)
                
                # Si c'est l'image principale et qu'une image principale existe déjà, la remplacer
                if is_main:
                    # Supprimer l'ancienne image principale si elle existe
                    ProductImage.objects.filter(product=product, is_main=True).delete()
                
                # Créer l'entrée dans la base de données
                product_image = ProductImage.objects.create(
                    product=product,
                    image=relative_path,
                    is_main=is_main,
                    order=0 if is_main else ProductImage.objects.filter(product=product).count()
                )
                
                return True, f"Image {'principale' if is_main else 'ajoutée'}: {filename}"
                
            except Exception as e:
                return False, f"Erreur lors de l'import de {os.path.basename(source_path)}: {e}"
        
        def process_product_folder(product_folder_path):
            """Traite un dossier de produit
            Structure attendue:
            - Dossier produit (nom du produit)
              - Sous-dossier référence (numéro de référence du produit)
                - Image/ (contient l'image principale)
                - Menu/ (contient les images de la galerie)
            """
            folder_name = os.path.basename(product_folder_path)
            logs = []
            
            # Chercher le sous-dossier de référence (premier sous-dossier)
            reference_folders = [d for d in os.listdir(product_folder_path) 
                                if os.path.isdir(os.path.join(product_folder_path, d))]
            
            if not reference_folders:
                logs.append(f"[ATTENTION] Aucun sous-dossier trouve")
                return {
                    'status': 'no_reference_folder',
                    'name': folder_name,
                    'logs': logs
                }
            
            # Prendre le premier sous-dossier (qui contient le numéro de référence)
            reference_folder = reference_folders[0]
            reference_path = os.path.join(product_folder_path, reference_folder)
            logs.append(f"[DOSSIER] Sous-dossier reference: {reference_folder}")
            
            # Extraire la reference du NOM DU SOUS-DOSSIER
            detected_ref = extract_reference_from_folder_name(reference_folder)
            if detected_ref:
                logs.append(f"[REF] Reference detectee: {detected_ref}")
            
            # Trouver le produit en utilisant le nom du sous-dossier de référence
            product, reference = find_product_by_reference(reference_folder)
            
            if not product:
                error_msg = f"[ATTENTION] Produit non trouve dans la base de donnees"
                logs.append(error_msg)
                logs.append(f"   Dossier: {folder_name}")
                logs.append(f"   Sous-dossier référence: {reference_folder}")
                if reference:
                    logs.append(f"   Référence recherchée: {reference}")
                return {
                    'status': 'not_found',
                    'name': folder_name,
                    'logs': logs
                }
            
            logs.append(f"[OK] Produit trouve: [{product.reference}] {product.name}")
            
            # Chercher les dossiers Image et Menu (insensible a la casse)
            image_folder = None
            menu_folder = None
            
            for item in os.listdir(reference_path):
                item_path = os.path.join(reference_path, item)
                if os.path.isdir(item_path):
                    item_lower = item.lower()
                    if item_lower == 'image':
                        image_folder = item_path
                    elif item_lower == 'menu':
                        menu_folder = item_path
            
            images_added = 0
            
            # Traiter l'image principale (dossier Image)
            if image_folder:
                image_files = get_image_files(image_folder)
                if image_files:
                    # Prendre la première image comme image principale
                    success, msg = copy_image_to_media(image_files[0], product, is_main=True)
                    if success:
                        images_added += 1
                        logs.append(f"[OK] {msg}")
                    else:
                        logs.append(f"[ERREUR] {msg}")
                else:
                    logs.append(f"[ATTENTION] Aucune image trouvee dans le dossier Image")
            else:
                logs.append(f"[ATTENTION] Dossier 'Image' non trouve")
            
            # Traiter les images supplementaires (dossier Menu)
            if menu_folder:
                menu_images = get_image_files(menu_folder)
                for image_path in menu_images:
                    success, msg = copy_image_to_media(image_path, product, is_main=False)
                    if success:
                        images_added += 1
                        logs.append(f"[OK] {msg}")
                    else:
                        logs.append(f"[ERREUR] {msg}")
            else:
                logs.append(f"[ATTENTION] Dossier 'Menu' non trouve")
            
            return {
                'status': 'success',
                'name': folder_name,
                'product': product,
                'images_count': images_added,
                'logs': logs
            }
        
        # Traiter tous les dossiers de produits
        try:
            product_folders = [
                os.path.join(images_path, d) 
                for d in os.listdir(images_path) 
                if os.path.isdir(os.path.join(images_path, d))
            ]
            
            # Statistiques
            stats = {
                'total': len(product_folders),
                'success': 0,
                'not_found': 0,
                'errors': 0,
                'total_images': 0
            }
            
            not_found_products = []
            detailed_logs = []
            
            # Traiter chaque dossier
            for product_folder in product_folders:
                try:
                    result = process_product_folder(product_folder)
                    
                    if result['status'] == 'success':
                        stats['success'] += 1
                        stats['total_images'] += result['images_count']
                    elif result['status'] == 'not_found':
                        stats['not_found'] += 1
                        not_found_products.append(result['name'])
                    else:
                        stats['errors'] += 1
                    
                    detailed_logs.extend(result['logs'])
                    
                except Exception as e:
                    stats['errors'] += 1
                    detailed_logs.append(f"[ERREUR] Erreur inattendue pour {os.path.basename(product_folder)}: {e}")
            
            # Message de succes
            success_msg = f"[OK] Importation terminee!\n"
            success_msg += f"- {stats['success']}/{stats['total']} produits traites avec succes\n"
            success_msg += f"- {stats['total_images']} images importees\n"
            success_msg += f"- {stats['not_found']} produits non trouves en base\n"
            success_msg += f"- {stats['errors']} erreurs"
            
            messages.success(request, success_msg)
            
            # Afficher les produits non trouves
            if not_found_products:
                warning_msg = f"[ATTENTION] Produits non trouves ({len(not_found_products)}):\n"
                warning_msg += "\n".join([f"- {name}" for name in not_found_products[:10]])
                if len(not_found_products) > 10:
                    warning_msg += f"\n... et {len(not_found_products) - 10} autres"
                messages.warning(request, warning_msg)
            
            # Sauvegarder les logs detailles dans la session pour affichage
            request.session['import_logs'] = detailed_logs
            
        except Exception as e:
            messages.error(request, f"[ERREUR] Erreur lors de l'importation: {str(e)}")
        
        return redirect('admin_panel:product_images_import')
    
    # Récupérer les logs de la dernière importation si disponibles
    import_logs = request.session.pop('import_logs', None)
    
    # Statistiques actuelles
    stats = {
        'products': Product.objects.count(),
        'products_with_images': Product.objects.filter(images__isnull=False).distinct().count(),
        'total_images': ProductImage.objects.count(),
    }
    
    return render(request, 'admin_panel/product_images_import.html', {
        'stats': stats,
        'import_logs': import_logs
    })


@login_required
def export_products_excel(request):
    """Exporte les produits vers un fichier Excel avec le format attendu pour l'import d'images"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"
    
    # Définir les colonnes (même format que data_product.xlsx)
    columns = [
        'Référence *',
        'Nom du produit *',
        'Catégorie *',
        'Sous-catégorie *',
        'Marque',
        'Type',
        'Collection',
        'Prix (DH) *',
        'Prix Promo (DH)',
        'Quantité *',
        'Statut',
        'Description *',
        'Caractéristiques',
        'Garantie',
        'Poids (kg)',
        'Meta Titre SEO',
        'Meta Description SEO',
        'Best Seller',
        'En vedette',
        'Nouveau'
    ]
    
    # Styles pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Écrire les en-têtes
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Récupérer les produits avec les relations
    products = Product.objects.select_related(
        'category', 'subcategory', 'brand', 'type', 'collection'
    ).order_by('reference')
    
    # Mapping des statuts
    status_mapping = {
        'in_stock': 'En Stock',
        'out_of_stock': 'Rupture de Stock',
        'preorder': 'Précommande',
        'discontinued': 'Discontinué',
    }
    
    # Écrire les données des produits
    for row_num, product in enumerate(products, 2):
        row_data = [
            product.reference,
            product.name,
            product.category.name if product.category else '',
            product.subcategory.name if product.subcategory else '',
            product.brand.name if product.brand else (product.brand_text or ''),
            product.type.name if product.type else '',
            product.collection.name if product.collection else '',
            float(product.price) if product.price else '',
            float(product.discount_price) if product.discount_price else '',
            product.quantity,
            status_mapping.get(product.status, product.status),
            product.description or '',
            product.caracteristiques or '',
            product.warranty or '',
            float(product.weight) if product.weight else '',
            product.meta_title or '',
            product.meta_description or '',
            'OUI' if product.is_bestseller else 'NON',
            'OUI' if product.is_featured else 'NON',
            'OUI' if product.is_new else 'NON'
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 50, 20, 25, 20, 20, 20, 12, 15, 10, 18, 60, 50, 20, 12, 50, 80, 12, 12, 10]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nom du fichier avec la date
    filename = f"produits_goback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le classeur dans la réponse
    wb.save(response)
    
    return response