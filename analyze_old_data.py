"""
Script pour analyser la structure du fichier old_data.xlsx
"""
import os
import sys
import django
import openpyxl
from pathlib import Path

# Configuration Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

def analyze_excel():
    """Analyse le fichier Excel et affiche sa structure"""
    excel_file = Path(r'C:\Users\MSI\Desktop\goback\old_data.xlsx')
    
    if not excel_file.exists():
        print(f"❌ Fichier {excel_file} introuvable!")
        print(f"📁 Répertoire actuel: {os.getcwd()}")
        return
    
    print(f"✅ Fichier trouvé: {excel_file}")
    print(f"📊 Analyse en cours...\n")
    
    # Charger le fichier Excel
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"📋 Feuilles disponibles: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"📄 FEUILLE: {sheet_name}")
        print(f"{'='*80}")
        
        # Nombre de lignes
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"📏 Dimensions: {max_row} lignes × {max_col} colonnes\n")
        
        # En-têtes (première ligne)
        headers = []
        print("📌 COLONNES:")
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
            print(f"  {col}. {cell_value}")
        
        # Exemples de données (lignes 2-4)
        print(f"\n📝 EXEMPLES DE DONNÉES (3 premières lignes):")
        for row in range(2, min(5, max_row + 1)):
            print(f"\n  --- Ligne {row} ---")
            for col, header in enumerate(headers, start=1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # Limiter l'affichage à 100 caractères
                    value_str = str(cell_value)
                    if len(value_str) > 100:
                        value_str = value_str[:100] + "..."
                    print(f"    {header}: {value_str}")
        
        # Statistiques
        print(f"\n📈 STATISTIQUES:")
        
        # Compter les valeurs non vides par colonne
        for col, header in enumerate(headers, start=1):
            non_empty = 0
            unique_values = set()
            for row in range(2, max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    non_empty += 1
                    # Pour certaines colonnes, compter les valeurs uniques
                    if header in ['Marque', 'Catégorie', 'Sous-catégorie', 'Type']:
                        unique_values.add(str(cell_value))
            
            print(f"  {header}: {non_empty}/{max_row-1} valeurs")
            if unique_values:
                print(f"    → {len(unique_values)} valeurs uniques: {list(unique_values)[:10]}")
    
    wb.close()

if __name__ == '__main__':
    analyze_excel()
