"""
Script d'importation complet depuis old_data.xlsx
Mapping:
- 1er niveau → SubCategory
- 2eme niveau → Type  
- Category → Category (garder la catégorie existante)
- marque → Brand
- reference, nom, description, prix, etc. → Product
- directory_path → copie des images (pic/ pour main, pics/ pour gallery)
"""
import os
import sys
import django
import openpyxl
import re
import shutil
from pathlib import Path
from django.utils.text import slugify
from django.core.files import File

# Configuration Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from shop.models import Category, SubCategory, Type, Brand, Product, ProductImage

# Chemin des médias Django
MEDIA_ROOT = Path(r'C:\Users\MSI\Desktop\goback\goback_backend\media')
PRODUCTS_IMAGES_DIR = MEDIA_ROOT / 'products'
PRODUCTS_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

def parse_description(description_text):
    """
    Parse la description pour extraire:
    - Description principale
    - Caractéristiques techniques
    - Dimensions
    """
    if not description_text:
        return "", "", ""
    
    desc = str(description_text)
    
    # Extraire les dimensions
    dimensions = ""
    dim_match = re.search(r'Dimensions?[:\s]*([^\n]+)', desc, re.IGNORECASE)
    if dim_match:
        dimensions = dim_match.group(1).strip()
    
    # Séparer description et caractéristiques
    parts = re.split(r'Caractéristiques techniques|Description\s*:', desc, flags=re.IGNORECASE)
    
    description = ""
    specifications = ""
    
    if len(parts) >= 2:
        description = parts[0].strip()
        specifications = '\n'.join(parts[1:]).strip()
    else:
        description = desc.strip()
    
    return description, specifications, dimensions


def copy_product_images(directory_path, reference):
    """
    Copie les images depuis le dossier source vers media/products/
    Retourne: (main_image_path, list_of_additional_images_paths)
    """
    if not directory_path or not os.path.exists(directory_path):
        print(f"  ⚠️  Dossier introuvable: {directory_path}")
        return None, []
    
    source_dir = Path(directory_path)
    main_image = None
    additional_images = []
    
    # Dossier de destination
    dest_dir = PRODUCTS_IMAGES_DIR / reference
    dest_dir.mkdir(parents=True, exist_ok=True)
    
    # Chercher dans le dossier 'pic' pour l'image principale
    pic_dir = source_dir / 'pic'
    if pic_dir.exists():
        for img_file in pic_dir.glob('*'):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp']:
                # Copier l'image
                dest_file = dest_dir / f"main{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                main_image = f"products/{reference}/main{img_file.suffix}"
                print(f"  ✅ Image principale copiée: {main_image}")
                break
    
    # Chercher dans le dossier 'pics' pour les images additionnelles
    pics_dir = source_dir / 'pics'
    if pics_dir.exists():
        for idx, img_file in enumerate(pics_dir.glob('*'), start=1):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp']:
                # Copier l'image
                dest_file = dest_dir / f"gallery_{idx}{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                img_path = f"products/{reference}/gallery_{idx}{img_file.suffix}"
                additional_images.append(img_path)
                print(f"  ✅ Image {idx} copiée: {img_path}")
    
    # Si pas d'image principale dans 'pic', chercher dans le dossier racine
    if not main_image:
        for img_file in source_dir.glob('*'):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp'] and img_file.is_file():
                dest_file = dest_dir / f"main{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                main_image = f"products/{reference}/main{img_file.suffix}"
                print(f"  ✅ Image principale copiée (racine): {main_image}")
                break
    
    return main_image, additional_images


def import_data():
    """Importation complète des données"""
    excel_file = Path(r'C:\Users\MSI\Desktop\goback\old_data.xlsx')
    
    if not excel_file.exists():
        print(f"❌ Fichier {excel_file} introuvable!")
        return
    
    print(f"✅ Début de l'importation depuis {excel_file}\n")
    
    # Charger Excel
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['example_imports_produits']
    
    # Statistiques
    stats = {
        'categories': 0,
        'subcategories': 0,
        'types': 0,
        'brands': 0,
        'products': 0,
        'images': 0,
        'errors': 0
    }
    
    # Cache pour éviter les duplications
    categories_cache = {}
    subcategories_cache = {}
    types_cache = {}
    brands_cache = {}
    
    print("🔄 Traitement des lignes...\n")
    
    for row_num in range(2, ws.max_row + 1):
        try:
            # Lire les données
            niveau_1 = ws.cell(row=row_num, column=1).value  # SubCategory
            niveau_2 = ws.cell(row=row_num, column=2).value  # Type
            category_name = ws.cell(row=row_num, column=3).value  # Category
            marque = ws.cell(row=row_num, column=4).value  # Brand
            reference = ws.cell(row=row_num, column=5).value
            nom = ws.cell(row=row_num, column=6).value
            description_raw = ws.cell(row=row_num, column=7).value
            prix = ws.cell(row=row_num, column=8).value
            prix_regulier = ws.cell(row=row_num, column=9).value
            qte = ws.cell(row=row_num, column=10).value
            directory_path = ws.cell(row=row_num, column=11).value
            coloris = ws.cell(row=row_num, column=12).value
            
            if not reference or not nom:
                continue
            
            print(f"📦 Ligne {row_num}: {nom} (Réf: {reference})")
            
            # 1. Créer/Récupérer la catégorie
            if category_name and category_name not in categories_cache:
                category, created = Category.objects.get_or_create(
                    name=str(category_name).strip(),
                    defaults={
                        'slug': slugify(category_name),
                        'description': f'Catégorie {category_name}',
                        'order': len(categories_cache) + 1
                    }
                )
                categories_cache[category_name] = category
                if created:
                    stats['categories'] += 1
                    print(f"  ✅ Catégorie créée: {category_name}")
            
            category = categories_cache.get(category_name)
            
            # 2. Créer/Récupérer la sous-catégorie (niveau 1)
            if niveau_1 and niveau_1 not in subcategories_cache:
                subcategory, created = SubCategory.objects.get_or_create(
                    name=str(niveau_1).strip(),
                    defaults={
                        'slug': slugify(niveau_1),
                        'category': category if category else Category.objects.first(),
                        'description': f'Collection {niveau_1}',
                        'order': len(subcategories_cache) + 1
                    }
                )
                subcategories_cache[niveau_1] = subcategory
                if created:
                    stats['subcategories'] += 1
                    print(f"  ✅ Sous-catégorie créée: {niveau_1}")
            
            subcategory = subcategories_cache.get(niveau_1)
            
            # 3. Créer/Récupérer le type (niveau 2)
            type_obj = None
            if niveau_2:
                type_key = f"{niveau_1}_{niveau_2}"
                if type_key not in types_cache:
                    type_obj, created = Type.objects.get_or_create(
                        name=str(niveau_2).strip(),
                        subcategory=subcategory,
                        defaults={
                            'slug': slugify(f"{niveau_1}-{niveau_2}"),
                            'description': f'Type {niveau_2}',
                            'order': len(types_cache) + 1
                        }
                    )
                    types_cache[type_key] = type_obj
                    if created:
                        stats['types'] += 1
                        print(f"  ✅ Type créé: {niveau_2}")
                else:
                    type_obj = types_cache[type_key]
            
            # 4. Créer/Récupérer la marque
            brand = None
            if marque:
                if marque not in brands_cache:
                    brand, created = Brand.objects.get_or_create(
                        name=str(marque).strip(),
                        defaults={
                            'slug': slugify(marque),
                            'description': f'Marque {marque}'
                        }
                    )
                    brands_cache[marque] = brand
                    if created:
                        stats['brands'] += 1
                        print(f"  ✅ Marque créée: {marque}")
                else:
                    brand = brands_cache[marque]
            
            # 5. Parser la description
            description, specifications, dimensions = parse_description(description_raw)
            
            # 6. Copier les images
            main_image, additional_images = copy_product_images(directory_path, reference)
            
            # 7. Créer le produit
            price_value = float(prix_regulier) if prix_regulier else (float(prix) if prix else 100.0)
            discount_price_value = float(prix) if prix and prix_regulier and float(prix) < float(prix_regulier) else None
            quantity_value = int(qte) if qte else 10
            
            product, created = Product.objects.update_or_create(
                reference=str(reference).strip(),
                defaults={
                    'name': str(nom).strip(),
                    'slug': slugify(f"{nom}-{reference}"),
                    'description': description[:2000] if description else f"Produit {nom}",
                    'caracteristiques': specifications[:1000] if specifications else "",
                    'price': str(price_value),
                    'discount_price': str(discount_price_value) if discount_price_value else None,
                    'quantity': quantity_value,
                    'status': 'in_stock' if quantity_value > 0 else 'out_of_stock',
                    'category': category,
                    'subcategory': subcategory,
                    'type': type_obj,
                    'brand': brand,
                    'main_image': main_image,
                    'warranty': dimensions if dimensions else "",
                    'is_featured': False,
                    'is_new': row_num <= 50  # Les 50 premiers sont nouveaux
                }
            )
            
            if created:
                stats['products'] += 1
                print(f"  ✅ Produit créé: {nom}")
            else:
                print(f"  ♻️  Produit mis à jour: {nom}")
            
            # 8. Ajouter les images additionnelles
            if additional_images:
                # Supprimer les anciennes images
                ProductImage.objects.filter(product=product).delete()
                
                for img_path in additional_images:
                    ProductImage.objects.create(
                        product=product,
                        image=img_path,
                        order=additional_images.index(img_path) + 1
                    )
                    stats['images'] += 1
            
            print(f"  ✅ Produit traité avec succès!\n")
            
        except Exception as e:
            stats['errors'] += 1
            print(f"  ❌ Erreur ligne {row_num}: {str(e)}\n")
            continue
    
    wb.close()
    
    # Afficher les statistiques finales
    print("\n" + "="*80)
    print("📊 RÉSUMÉ DE L'IMPORTATION")
    print("="*80)
    print(f"✅ Catégories créées: {stats['categories']}")
    print(f"✅ Sous-catégories créées: {stats['subcategories']}")
    print(f"✅ Types créés: {stats['types']}")
    print(f"✅ Marques créées: {stats['brands']}")
    print(f"✅ Produits créés/mis à jour: {stats['products']}")
    print(f"🖼️  Images copiées: {stats['images']}")
    print(f"❌ Erreurs: {stats['errors']}")
    print("="*80)

if __name__ == '__main__':
    import_data()
