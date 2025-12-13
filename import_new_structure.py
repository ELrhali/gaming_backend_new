"""
Script d'importation avec nouvelle structure
1er niveau → Category
2eme niveau → SubCategory  
Category → Type
"""
import os
import sys
import django
import openpyxl
import re
import shutil
import random
from pathlib import Path
from django.utils.text import slugify

# Configuration Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from shop.models import Category, SubCategory, Type, Brand, Product, ProductImage

# Chemin des médias Django
MEDIA_ROOT = Path(r'C:\Users\MSI\Desktop\goback\goback_backend\media')
PRODUCTS_IMAGES_DIR = MEDIA_ROOT / 'products'
PRODUCTS_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

def parse_description(description_text):
    """Parse la description pour extraire dimensions et caractéristiques"""
    if not description_text:
        return "", ""
    
    desc = str(description_text)
    
    # Séparer description et caractéristiques
    parts = re.split(r'Caractéristiques techniques|Description\s*:', desc, flags=re.IGNORECASE)
    
    description = ""
    specifications = ""
    
    if len(parts) >= 2:
        description = parts[0].strip()
        specifications = '\n'.join(parts[1:]).strip()
    else:
        description = desc.strip()
    
    return description, specifications


def copy_product_images(directory_path, reference):
    """
    Copie les images depuis le dossier source
    Retourne: (main_image_path, list_of_additional_images_paths)
    """
    if not directory_path:
        return None, []
    
    # Nettoyer le chemin (supprimer les \n et espaces)
    directory_path = str(directory_path).strip()
    
    if not os.path.exists(directory_path):
        print(f"  ⚠️  Dossier introuvable: {directory_path}")
        return None, []
    
    source_dir = Path(directory_path)
    main_image = None
    additional_images = []
    
    # Dossier de destination
    dest_dir = PRODUCTS_IMAGES_DIR / reference
    dest_dir.mkdir(parents=True, exist_ok=True)
    
    # Chercher dans 'pic' pour l'image principale
    pic_dir = source_dir / 'pic'
    if pic_dir.exists():
        for img_file in pic_dir.glob('*'):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp']:
                dest_file = dest_dir / f"main{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                main_image = f"products/{reference}/main{img_file.suffix}"
                print(f"  [IMG] Image principale: {img_file.name}")
                break
    
    # Chercher dans 'pics' pour galerie
    pics_dir = source_dir / 'pics'
    if pics_dir.exists():
        for idx, img_file in enumerate(pics_dir.glob('*'), start=1):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp']:
                dest_file = dest_dir / f"gallery_{idx}{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                img_path = f"products/{reference}/gallery_{idx}{img_file.suffix}"
                additional_images.append(img_path)
                if idx <= 2:  # Afficher seulement les 2 premières
                    print(f"  [IMG] Image {idx}: {img_file.name}")
    
    # Si pas d'image principale dans 'pic', chercher dans racine
    if not main_image:
        for img_file in source_dir.glob('*'):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.webp'] and img_file.is_file():
                dest_file = dest_dir / f"main{img_file.suffix}"
                shutil.copy2(img_file, dest_file)
                main_image = f"products/{reference}/main{img_file.suffix}"
                print(f"  [IMG] Image principale (racine): {img_file.name}")
                break
    
    return main_image, additional_images


def generate_random_price():
    """Génère un prix aléatoire entre 50 et 2000 DH"""
    return round(random.uniform(50, 2000), 2)


def generate_random_quantity():
    """Génère une quantité aléatoire entre 0 et 100"""
    return random.randint(0, 100)


def import_data():
    """Importation avec nouvelle structure"""
    excel_file = Path(r'C:\Users\MSI\Desktop\goback\old_data.xlsx')
    
    if not excel_file.exists():
        print(f"❌ Fichier {excel_file} introuvable!")
        return
    
    print(f"[OK] Debut de l'importation depuis {excel_file}\n")
    print("Structure:")
    print("  - 1er niveau -> Category")
    print("  - 2eme niveau -> SubCategory")
    print("  - Category -> Type\n")
    
    # Charger Excel
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['example_imports_produits']
    
    # Statistiques
    stats = {
        'categories': 0,
        'subcategories': 0,
        'types': 0,
        'brands': 0,
        'products': 0,
        'images_main': 0,
        'images_gallery': 0,
        'errors': 0
    }
    
    # Cache
    categories_cache = {}
    subcategories_cache = {}
    types_cache = {}
    brands_cache = {}
    
    print("🔄 Traitement des produits...\n")
    
    for row_num in range(2, ws.max_row + 1):
        try:
            # Lire les données avec nouvelle structure
            category_name = ws.cell(row=row_num, column=1).value  # 1er niveau → Category
            subcategory_name = ws.cell(row=row_num, column=2).value  # 2eme niveau → SubCategory
            type_name = ws.cell(row=row_num, column=3).value  # Category → Type
            marque = ws.cell(row=row_num, column=4).value
            reference = ws.cell(row=row_num, column=5).value
            nom = ws.cell(row=row_num, column=6).value
            description_raw = ws.cell(row=row_num, column=7).value
            prix = ws.cell(row=row_num, column=8).value
            prix_regulier = ws.cell(row=row_num, column=9).value
            qte = ws.cell(row=row_num, column=10).value
            directory_path = ws.cell(row=row_num, column=11).value
            coloris = ws.cell(row=row_num, column=12).value
            
            if not reference or not nom:
                continue
            
            print(f"[{row_num-1}/{ws.max_row-1}] {nom[:50]}... (Ref: {reference})")
            
            # 1. Créer/Récupérer la CATÉGORIE (1er niveau)
            category = None
            if category_name:
                if category_name not in categories_cache:
                    category, created = Category.objects.get_or_create(
                        name=str(category_name).strip(),
                        defaults={
                            'slug': slugify(category_name),
                            'description': f'Catégorie {category_name}',
                            'order': len(categories_cache) + 1
                        }
                    )
                    categories_cache[category_name] = category
                    if created:
                        stats['categories'] += 1
                        print(f"  [+] Categorie: {category_name}")
                else:
                    category = categories_cache[category_name]
            
            # 2. Créer/Récupérer la SOUS-CATÉGORIE (2eme niveau)
            subcategory = None
            if subcategory_name and category:
                subcategory_key = f"{category_name}_{subcategory_name}"
                if subcategory_key not in subcategories_cache:
                    subcategory, created = SubCategory.objects.get_or_create(
                        name=str(subcategory_name).strip(),
                        category=category,
                        defaults={
                            'slug': slugify(f"{category_name}-{subcategory_name}"),
                            'description': f'{subcategory_name}',
                            'order': len(subcategories_cache) + 1
                        }
                    )
                    subcategories_cache[subcategory_key] = subcategory
                    if created:
                        stats['subcategories'] += 1
                        print(f"  [+] Sous-categorie: {subcategory_name}")
                else:
                    subcategory = subcategories_cache[subcategory_key]
            
            # 3. Créer/Récupérer le TYPE (colonne Category)
            type_obj = None
            if type_name and subcategory:
                # Clé unique basée sur catégorie + sous-catégorie + type
                type_key = f"{category_name}_{subcategory_name}_{type_name}"
                if type_key not in types_cache:
                    # Slug unique incluant la catégorie pour éviter les doublons
                    unique_slug = slugify(f"{category_name}-{subcategory_name}-{type_name}")
                    type_obj, created = Type.objects.get_or_create(
                        name=str(type_name).strip(),
                        subcategory=subcategory,
                        defaults={
                            'slug': unique_slug,
                            'description': f'Type {type_name}',
                            'order': len(types_cache) + 1
                        }
                    )
                    types_cache[type_key] = type_obj
                    if created:
                        stats['types'] += 1
                        print(f"  [+] Type: {type_name}")
                else:
                    type_obj = types_cache[type_key]
            
            # 4. Créer/Récupérer la MARQUE
            brand = None
            if marque:
                if marque not in brands_cache:
                    brand, created = Brand.objects.get_or_create(
                        name=str(marque).strip(),
                        defaults={
                            'slug': slugify(marque),
                            'description': f'Marque {marque}'
                        }
                    )
                    brands_cache[marque] = brand
                    if created:
                        stats['brands'] += 1
                        print(f"  [+] Marque: {marque}")
                else:
                    brand = brands_cache[marque]
            
            # 5. Parser la description
            description, specifications = parse_description(description_raw)
            
            # 6. Générer prix et quantité aléatoires si vides
            if prix and prix_regulier:
                price_value = float(prix_regulier)
                discount_price_value = float(prix) if float(prix) < float(prix_regulier) else None
            else:
                price_value = generate_random_price()
                discount_price_value = generate_random_price() * 0.85  # 15% de réduction
            
            quantity_value = int(qte) if qte else generate_random_quantity()
            
            # 7. Copier les images
            main_image, additional_images = copy_product_images(directory_path, str(reference).strip())
            
            if main_image:
                stats['images_main'] += 1
            stats['images_gallery'] += len(additional_images)
            
            # 8. Créer le PRODUIT
            product, created = Product.objects.update_or_create(
                reference=str(reference).strip(),
                defaults={
                    'name': str(nom).strip(),
                    'slug': slugify(f"{nom}-{reference}"),
                    'description': description[:2000] if description else f"Produit {nom}",
                    'caracteristiques': specifications[:1000] if specifications else "",
                    'price': str(price_value),
                    'discount_price': str(discount_price_value) if discount_price_value else None,
                    'quantity': quantity_value,
                    'status': 'in_stock' if quantity_value > 0 else 'out_of_stock',
                    'category': category,
                    'subcategory': subcategory,
                    'type': type_obj,
                    'brand': brand,
                    'main_image': main_image,
                    'is_featured': random.choice([True, False]),
                    'is_new': row_num <= 30
                }
            )
            
            # 9. Ajouter les images additionnelles
            if additional_images:
                ProductImage.objects.filter(product=product).delete()
                for img_path in additional_images:
                    ProductImage.objects.create(
                        product=product,
                        image=img_path,
                        order=additional_images.index(img_path) + 1
                    )
            
            if created:
                stats['products'] += 1
                print(f"  [OK] Produit cree | Prix: {price_value} DH | Stock: {quantity_value}\n")
            else:
                print(f"  [UPD] Produit mis a jour\n")
            
        except Exception as e:
            stats['errors'] += 1
            print(f"  [ERR] Erreur ligne {row_num}: {str(e)}\n")
            continue
    
    wb.close()
    
    # Afficher les statistiques finales
    print("\n" + "="*80)
    print("RESUME DE L'IMPORTATION")
    print("="*80)
    print(f"[+] Categories creees       : {stats['categories']}")
    print(f"[+] Sous-categories creees  : {stats['subcategories']}")
    print(f"[+] Types crees             : {stats['types']}")
    print(f"[+] Marques creees          : {stats['brands']}")
    print(f"[+] Produits crees          : {stats['products']}")
    print(f"[IMG] Images principales    : {stats['images_main']}")
    print(f"[IMG] Images galerie        : {stats['images_gallery']}")
    print(f"[ERR] Erreurs               : {stats['errors']}")
    print("="*80)

if __name__ == '__main__':
    import_data()
