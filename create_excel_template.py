#!/usr/bin/env python
"""
Script pour créer un fichier Excel exemple avec des produits
Ce fichier peut servir de template pour importer des produits
"""
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Créer un nouveau classeur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Produits"

# Définir les en-têtes avec descriptions
headers = [
    ("reference", "Référence unique du produit (ex: RTX4090-001)"),
    ("name", "Nom complet du produit"),
    ("category", "Nom de la catégorie (ex: Composants, PC, Périphériques)"),
    ("subcategory", "Nom de la sous-catégorie (ex: Cartes Graphiques, Cartes Mères)"),
    ("brand", "Nom de la marque (ex: ASUS, MSI, Gigabyte)"),
    ("type", "Modèle/Type (ex: ROG Strix, Gaming X, TUF)"),
    ("collection", "Collection (optionnel, ex: Gaming, Pro)"),
    ("price", "Prix normal (en DH, ex: 15999.99)"),
    ("discount_price", "Prix promo (optionnel, en DH, ex: 14999.99)"),
    ("quantity", "Quantité en stock (nombre entier, ex: 10)"),
    ("status", "Statut (in_stock, out_of_stock, preorder, discontinued)"),
    ("description", "Description longue du produit"),
    ("caracteristiques", "Caractéristiques principales (séparées par des retours à la ligne)"),
    ("warranty", "Garantie (ex: 2 ans, 3 ans constructeur)"),
    ("weight", "Poids en kg (optionnel, ex: 1.5)"),
    ("is_bestseller", "Best Seller ? (OUI ou NON)"),
    ("is_featured", "En vedette ? (OUI ou NON)"),
    ("is_new", "Nouveau ? (OUI ou NON)"),
    ("meta_title", "Titre SEO (optionnel)"),
    ("meta_description", "Description SEO (optionnel)"),
]

# Ajouter les en-têtes
for col_idx, (header, description) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter la description en ligne 2
    desc_cell = ws.cell(row=2, column=col_idx)
    desc_cell.value = description
    desc_cell.font = Font(italic=True, size=9)
    desc_cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

# Ajuster la hauteur des lignes d'en-tête
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 40

# Données d'exemple
products_data = [
    {
        "reference": "RTX4090-ASUS-001",
        "name": "ASUS ROG Strix GeForce RTX 4090 OC Edition 24GB GDDR6X",
        "category": "Composants",
        "subcategory": "Cartes Graphiques",
        "brand": "ASUS",
        "type": "ROG Strix",
        "collection": "Gaming",
        "price": "25999.00",
        "discount_price": "24499.00",
        "quantity": "5",
        "status": "in_stock",
        "description": "La carte graphique ASUS ROG Strix GeForce RTX 4090 OC Edition offre des performances exceptionnelles pour le gaming 4K et la création de contenu. Équipée de 24GB de mémoire GDDR6X et du système de refroidissement Axial-tech, elle garantit des températures optimales même sous charge intensive.",
        "caracteristiques": "GPU: NVIDIA GeForce RTX 4090\nMémoire: 24GB GDDR6X\nBus mémoire: 384-bit\nClock Boost: 2640 MHz\nPorts: 3x DisplayPort 1.4a, 2x HDMI 2.1\nRefroidissement: Axial-tech avec 3 ventilateurs\nRGB: Aura Sync compatible",
        "warranty": "3 ans constructeur",
        "weight": "2.5",
        "is_bestseller": "OUI",
        "is_featured": "OUI",
        "is_new": "NON",
        "meta_title": "ASUS ROG Strix RTX 4090 OC 24GB - Meilleur Prix Maroc",
        "meta_description": "Achetez la carte graphique ASUS ROG Strix GeForce RTX 4090 OC Edition 24GB au meilleur prix au Maroc. Livraison rapide et garantie 3 ans.",
    },
    {
        "reference": "MB-Z790-MSI-001",
        "name": "MSI MAG Z790 TOMAHAWK WIFI DDR5 - Carte Mère Gaming ATX",
        "category": "Composants",
        "subcategory": "Cartes Mères",
        "brand": "MSI",
        "type": "MAG",
        "collection": "Gaming",
        "price": "3899.00",
        "discount_price": "",
        "quantity": "12",
        "status": "in_stock",
        "description": "La carte mère MSI MAG Z790 TOMAHAWK WIFI offre une base solide pour les processeurs Intel de 13e et 14e génération. Avec support DDR5, PCIe 5.0 et Wi-Fi 6E intégré, elle est parfaite pour les configurations gaming haut de gamme.",
        "caracteristiques": "Chipset: Intel Z790\nSocket: LGA 1700\nMémoire: 4x DDR5 jusqu'à 7200+ MHz\nSlots PCIe: 1x PCIe 5.0 x16, 2x PCIe 4.0 x16\nStockage: 4x M.2, 6x SATA\nRéseau: 2.5G LAN + Wi-Fi 6E\nUSB: 1x USB 3.2 Gen 2x2 Type-C, plusieurs USB 3.2 Gen 2/Gen 1",
        "warranty": "3 ans constructeur",
        "weight": "1.2",
        "is_bestseller": "OUI",
        "is_featured": "NON",
        "is_new": "NON",
        "meta_title": "MSI MAG Z790 TOMAHAWK WIFI DDR5 - Carte Mère Gaming",
        "meta_description": "Carte mère MSI MAG Z790 TOMAHAWK WIFI DDR5 pour processeurs Intel Gen 13/14. Support DDR5, PCIe 5.0, Wi-Fi 6E. Disponible au Maroc.",
    },
    {
        "reference": "RAM-DDR5-GSKILL-001",
        "name": "G.SKILL Trident Z5 RGB 32GB (2x16GB) DDR5 6000MHz CL30",
        "category": "Composants",
        "subcategory": "Mémoire RAM",
        "brand": "G.SKILL",
        "type": "Trident Z5",
        "collection": "RGB",
        "price": "1899.00",
        "discount_price": "1799.00",
        "quantity": "20",
        "status": "in_stock",
        "description": "Kit de mémoire DDR5 G.SKILL Trident Z5 RGB 32GB (2x16GB) cadencé à 6000MHz avec des timings CL30. Idéal pour les configurations gaming et workstation haute performance avec éclairage RGB personnalisable.",
        "caracteristiques": "Capacité: 32GB (2x16GB)\nType: DDR5\nFréquence: 6000MHz\nLatence: CL30-38-38-96\nVoltage: 1.35V\nRGB: Oui, personnalisable\nCompatibilité: Intel XMP 3.0\nProfil: Optimisé pour Intel 12e/13e/14e gen",
        "warranty": "Garantie à vie",
        "weight": "0.3",
        "is_bestseller": "OUI",
        "is_featured": "OUI",
        "is_new": "NON",
        "meta_title": "G.SKILL Trident Z5 RGB 32GB DDR5 6000MHz - RAM Gaming",
        "meta_description": "Kit RAM G.SKILL Trident Z5 RGB 32GB DDR5 6000MHz CL30. Performance maximale pour gaming et création. Disponible au Maroc avec garantie.",
    },
    {
        "reference": "SSD-2TB-SAMSUNG-001",
        "name": "Samsung 990 PRO 2TB NVMe M.2 SSD PCIe 4.0",
        "category": "Composants",
        "subcategory": "Stockage",
        "brand": "Samsung",
        "type": "990 PRO",
        "collection": "",
        "price": "2499.00",
        "discount_price": "",
        "quantity": "15",
        "status": "in_stock",
        "description": "Le SSD Samsung 990 PRO offre des vitesses exceptionnelles jusqu'à 7450 MB/s en lecture et 6900 MB/s en écriture. Idéal pour le gaming, la création de contenu et les applications professionnelles nécessitant des débits élevés.",
        "caracteristiques": "Capacité: 2TB\nInterface: PCIe 4.0 x4, NVMe 2.0\nFormat: M.2 2280\nVitesse lecture: jusqu'à 7450 MB/s\nVitesse écriture: jusqu'à 6900 MB/s\nIOPS lecture: 1400K\nIOPS écriture: 1550K\nTechnologie: V-NAND 3-bit MLC\nGarantie: 5 ans ou 1200 TBW",
        "warranty": "5 ans constructeur",
        "weight": "0.01",
        "is_bestseller": "OUI",
        "is_featured": "NON",
        "is_new": "NON",
        "meta_title": "Samsung 990 PRO 2TB NVMe SSD PCIe 4.0 - Stockage Rapide",
        "meta_description": "SSD Samsung 990 PRO 2TB NVMe PCIe 4.0 - Vitesses jusqu'à 7450 MB/s. Le meilleur SSD pour gaming et création au Maroc.",
    },
    {
        "reference": "ECRAN-27-ASUS-001",
        "name": "ASUS ROG Swift PG27AQDM 27\" OLED QHD 240Hz Gaming Monitor",
        "category": "Périphériques",
        "subcategory": "Écrans",
        "brand": "ASUS",
        "type": "ROG Swift",
        "collection": "Gaming",
        "price": "12999.00",
        "discount_price": "11999.00",
        "quantity": "3",
        "status": "in_stock",
        "description": "L'écran gaming ASUS ROG Swift PG27AQDM dispose d'une dalle OLED QHD de 27 pouces avec un taux de rafraîchissement de 240Hz. Temps de réponse de 0.03ms, G-SYNC compatible et couverture DCI-P3 99% pour une expérience visuelle inégalée.",
        "caracteristiques": "Taille: 27 pouces\nRésolution: 2560 x 1440 (QHD)\nDalle: OLED\nTaux de rafraîchissement: 240Hz\nTemps de réponse: 0.03ms (GTG)\nLuminosité: 450 cd/m² (pic 1000 cd/m²)\nContraste: 1,500,000:1\nCouverture couleur: 99% DCI-P3\nG-SYNC: Compatible\nPorts: 2x HDMI 2.0, 1x DisplayPort 1.4, USB Hub",
        "warranty": "3 ans dont 1 an anti-burn-in",
        "weight": "6.5",
        "is_bestseller": "OUI",
        "is_featured": "OUI",
        "is_new": "OUI",
        "meta_title": "ASUS ROG Swift PG27AQDM 27\" OLED 240Hz - Écran Gaming",
        "meta_description": "Écran gaming OLED ASUS ROG Swift PG27AQDM 27\" QHD 240Hz. Image parfaite et réactivité ultime pour le gaming compétitif au Maroc.",
    },
    {
        "reference": "KB-LOGI-001",
        "name": "Logitech G PRO X TKL LIGHTSPEED Clavier Gaming Sans Fil",
        "category": "Périphériques",
        "subcategory": "Claviers",
        "brand": "Logitech",
        "type": "G PRO",
        "collection": "Gaming",
        "price": "2299.00",
        "discount_price": "",
        "quantity": "8",
        "status": "in_stock",
        "description": "Le clavier gaming sans fil Logitech G PRO X TKL LIGHTSPEED est conçu pour les gamers professionnels. Format compact TKL, switches mécaniques GX au choix, technologie LIGHTSPEED pour une latence ultra-faible et RGB LIGHTSYNC.",
        "caracteristiques": "Format: TKL (Tenkeyless) - Sans pavé numérique\nConnexion: Sans fil LIGHTSPEED 2.4GHz + USB-C\nSwitches: GX Red/Blue/Brown (au choix)\nRGB: LIGHTSYNC RGB par touche\nAutonomie: jusqu'à 50h\nTaux de rapport: 1000Hz (1ms)\nConstruction: Aluminium brossé\nCompatibilité: Windows, macOS",
        "warranty": "2 ans constructeur",
        "weight": "0.78",
        "is_bestseller": "OUI",
        "is_featured": "NON",
        "is_new": "NON",
        "meta_title": "Logitech G PRO X TKL Sans Fil - Clavier Gaming Pro",
        "meta_description": "Clavier gaming Logitech G PRO X TKL LIGHTSPEED sans fil. Performance pro, format compact TKL, RGB personnalisable. Disponible au Maroc.",
    },
    {
        "reference": "MOUSE-RAZER-001",
        "name": "Razer Viper V3 Pro - Souris Gaming Sans Fil Ultra-Légère",
        "category": "Périphériques",
        "subcategory": "Souris",
        "brand": "Razer",
        "type": "Viper V3",
        "collection": "Pro",
        "price": "1899.00",
        "discount_price": "1699.00",
        "quantity": "10",
        "status": "in_stock",
        "description": "La Razer Viper V3 Pro est une souris gaming sans fil ultra-légère de seulement 54g. Équipée du capteur Focus Pro 30K, switches optiques Gen-3 et technologie HyperSpeed Wireless pour une performance sans compromis.",
        "caracteristiques": "Capteur: Focus Pro 30K DPI (jusqu'à 30,000 DPI)\nPoids: 54g\nConnexion: HyperSpeed Wireless + USB-C filaire\nSwitches: Optiques Gen-3 (90M de clics)\nAutonomie: jusqu'à 95h\nTaux de rapport: 1000Hz (1ms sans fil)\nRGB: Razer Chroma RGB\nBoutons: 5 programmables\nPatins: PTFE pour glisse optimale",
        "warranty": "2 ans constructeur",
        "weight": "0.054",
        "is_bestseller": "OUI",
        "is_featured": "OUI",
        "is_new": "OUI",
        "meta_title": "Razer Viper V3 Pro - Souris Gaming Sans Fil Ultra-Légère",
        "meta_description": "Souris gaming Razer Viper V3 Pro sans fil 54g. Capteur 30K DPI, switches optiques, 95h d'autonomie. Gaming pro au Maroc.",
    },
    {
        "reference": "PC-AMD-CUSTOM-001",
        "name": "PC Gamer AMD Ryzen 7 7800X3D RTX 4070 Ti 32GB DDR5 1TB SSD",
        "category": "PC",
        "subcategory": "PC Gaming",
        "brand": "",
        "type": "",
        "collection": "Gaming",
        "price": "18999.00",
        "discount_price": "",
        "quantity": "0",
        "status": "preorder",
        "description": "PC Gaming sur-mesure équipé d'un processeur AMD Ryzen 7 7800X3D, carte graphique RTX 4070 Ti, 32GB RAM DDR5 et SSD 1TB NVMe. Configuration optimale pour le gaming 1440p/4K et le streaming.",
        "caracteristiques": "Processeur: AMD Ryzen 7 7800X3D (8 cœurs, 16 threads, jusqu'à 5.0GHz)\nCarte Graphique: NVIDIA GeForce RTX 4070 Ti 12GB\nMémoire: 32GB DDR5 6000MHz (2x16GB)\nStockage: 1TB NVMe Gen 4 SSD\nCarte Mère: B650 ATX Wi-Fi\nAlimentation: 850W 80+ Gold Modulaire\nRefroidissement: AIO 240mm RGB\nBoîtier: Moyen Tour RGB Tempered Glass\nWindows 11 Pro préinstallé",
        "warranty": "2 ans pièces et main d'œuvre",
        "weight": "15.0",
        "is_bestseller": "NON",
        "is_featured": "OUI",
        "is_new": "OUI",
        "meta_title": "PC Gamer AMD Ryzen 7 7800X3D RTX 4070 Ti - Config Gaming",
        "meta_description": "PC Gaming AMD Ryzen 7 7800X3D + RTX 4070 Ti 32GB DDR5. Configuration haute performance pour gaming 1440p/4K. Assemblé au Maroc.",
    },
]

# Ajouter les données (commencer à la ligne 3)
for row_idx, product in enumerate(products_data, 3):
    for col_idx, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = product.get(header, "")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Ajuster la largeur des colonnes
column_widths = {
    'A': 20,  # reference
    'B': 50,  # name
    'C': 15,  # category
    'D': 20,  # subcategory
    'E': 15,  # brand
    'F': 20,  # type
    'G': 15,  # collection
    'H': 12,  # price
    'I': 12,  # discount_price
    'J': 10,  # quantity
    'K': 15,  # status
    'L': 60,  # description
    'M': 60,  # caracteristiques
    'N': 20,  # warranty
    'O': 10,  # weight
    'P': 12,  # is_bestseller
    'Q': 12,  # is_featured
    'R': 12,  # is_new
    'S': 40,  # meta_title
    'T': 50,  # meta_description
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Figer les deux premières lignes
ws.freeze_panes = 'A3'

# Créer une feuille d'instructions
instructions_sheet = wb.create_sheet("Instructions")
instructions_sheet.column_dimensions['A'].width = 100

instructions = [
    ("GUIDE D'UTILISATION - TEMPLATE PRODUITS", "title"),
    ("", "empty"),
    ("Ce fichier Excel contient des exemples de produits pour vous aider à comprendre le format requis.", "text"),
    ("", "empty"),
    ("COLONNES OBLIGATOIRES:", "subtitle"),
    ("• reference: Code unique du produit (ne peut pas être dupliqué)", "text"),
    ("• name: Nom complet du produit", "text"),
    ("• category: Nom exact de la catégorie (doit exister dans la base)", "text"),
    ("• subcategory: Nom exact de la sous-catégorie (doit exister dans la base)", "text"),
    ("• price: Prix en DH (format: 1999.99)", "text"),
    ("• quantity: Nombre en stock (entier)", "text"),
    ("• status: in_stock, out_of_stock, preorder, ou discontinued", "text"),
    ("• description: Description détaillée du produit", "text"),
    ("", "empty"),
    ("COLONNES OPTIONNELLES:", "subtitle"),
    ("• brand: Nom de la marque (doit exister dans la base)", "text"),
    ("• type: Modèle/Type du produit (doit exister dans la base)", "text"),
    ("• collection: Collection (optionnel)", "text"),
    ("• discount_price: Prix promotionnel (laisser vide si pas de promo)", "text"),
    ("• caracteristiques: Caractéristiques techniques (utiliser Alt+Entrée pour retour à la ligne)", "text"),
    ("• warranty: Garantie (ex: 2 ans, 3 ans constructeur)", "text"),
    ("• weight: Poids en kg (ex: 1.5)", "text"),
    ("• is_bestseller: OUI ou NON", "text"),
    ("• is_featured: OUI ou NON", "text"),
    ("• is_new: OUI ou NON", "text"),
    ("• meta_title: Titre SEO (optionnel)", "text"),
    ("• meta_description: Description SEO (optionnel)", "text"),
    ("", "empty"),
    ("IMPORTANT:", "subtitle"),
    ("1. Les noms de catégories, sous-catégories, marques et types doivent EXACTEMENT correspondre", "text"),
    ("   à ceux qui existent déjà dans votre base de données.", "text"),
    ("2. N'utilisez pas de guillemets dans les textes", "text"),
    ("3. Pour les retours à la ligne dans une cellule, utilisez Alt+Entrée", "text"),
    ("4. Les prix doivent utiliser le point comme séparateur décimal (ex: 1999.99)", "text"),
    ("5. Ce fichier est un EXEMPLE - vous pouvez ajouter autant de lignes que nécessaire", "text"),
    ("", "empty"),
    ("CATÉGORIES EXISTANTES (exemples):", "subtitle"),
    ("• Composants", "text"),
    ("• PC", "text"),
    ("• Périphériques", "text"),
    ("• Accessoires", "text"),
    ("", "empty"),
    ("STATUS POSSIBLES:", "subtitle"),
    ("• in_stock: En stock", "text"),
    ("• out_of_stock: Rupture de stock", "text"),
    ("• preorder: Précommande", "text"),
    ("• discontinued: Discontinué", "text"),
]

for row_idx, (text, style_type) in enumerate(instructions, 1):
    cell = instructions_sheet.cell(row=row_idx, column=1)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    if style_type == "title":
        cell.font = Font(bold=True, size=16, color="0066CC")
        cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        instructions_sheet.row_dimensions[row_idx].height = 25
    elif style_type == "subtitle":
        cell.font = Font(bold=True, size=12, color="0066CC")
        instructions_sheet.row_dimensions[row_idx].height = 20
    elif style_type == "text":
        cell.font = Font(size=10)
        instructions_sheet.row_dimensions[row_idx].height = 18
    elif style_type == "empty":
        instructions_sheet.row_dimensions[row_idx].height = 10

# Sauvegarder le fichier
output_file = "exemple_produits.xlsx"
wb.save(output_file)

print("=" * 70)
print("✅ FICHIER EXCEL CRÉÉ AVEC SUCCÈS!")
print("=" * 70)
print(f"\n📄 Nom du fichier: {output_file}")
print(f"📍 Emplacement: {os.path.abspath(output_file)}")
print(f"\n📊 Contenu:")
print(f"   • Feuille 'Produits': {len(products_data)} exemples de produits")
print(f"   • Feuille 'Instructions': Guide d'utilisation complet")
print(f"\n💡 Ce fichier contient des exemples réels de produits gaming")
print(f"   pour vous aider à comprendre le format et les données requises.")
print(f"\n🎯 Vous pouvez:")
print(f"   1. Modifier les exemples existants")
print(f"   2. Ajouter de nouvelles lignes pour vos produits")
print(f"   3. Utiliser ce fichier comme template pour créer votre catalogue")
print("\n" + "=" * 70)
